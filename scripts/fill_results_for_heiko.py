"""
Fill compliance columns in docs/results_for_heiko.xlsx from ppwr_audit_results.csv.

Matching: audit Supplier No. (e.g. 0100114879) -> Heiko Supplier ID (e.g. 100114879).
Values written: 1 = compliant (Yes), 0 = non-compliant (No), "no response" = N/A.

Compliance semantics come from dashboard_ppwr.resolve_check (single source of truth,
handles the yes/no inversion, PFAS pass-through, evidence checks, etc.).

Write strategy: surgical XML edit. We only rewrite the 4 compliance cells of the
matched rows inside xl/worksheets/sheet1.xml and repackage the zip. The ~99k data
rows and all other columns/formatting are copied byte-for-byte. This is fast and
cannot silently drop the master data (unlike a full openpyxl/pandas rewrite).
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import zipfile
from datetime import datetime
from typing import Dict, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

# Make the repo root importable when run as scripts/fill_results_for_heiko.py.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(BASE_DIR)
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from dashboard_ppwr import CHECK_ORDER, resolve_check  # noqa: E402

AUDIT_CSV = os.path.join(ROOT_DIR, "data", "ppwr_audit_results.csv")
HEIKO_XLSX = os.path.join(ROOT_DIR, "data", "results", "1st_2nd_wave_results_for_heiko.xlsx")

DATA_START_ROW = 4  # Excel row 4 = first data row (header on row 3)
SHEET_XML = "xl/worksheets/sheet1.xml"

# Excel column index per check (F=6 heavy metals, G=7 PFAS, H=8 SoC, I=9 SVHC).
COL_COMPLIANCE = {
    "Heavy metals": 6,
    "PFAS": 7,
    "SoC": 8,
    "SVHC": 9,
}


# --------------------------------------------------------------------------- #
# Compliance lookup                                                           #
# --------------------------------------------------------------------------- #
def audit_supplier_to_heiko_key(supplier_no: str) -> Optional[str]:
    text = str(supplier_no or "").strip()
    if not text:
        return None
    if text.isdigit() or text.lstrip("0").isdigit():
        return str(int(text))
    return None


def display_to_binary(display: str):
    if display == "Yes":
        return 1
    if display == "No":
        return 0
    return "no response"


def normalize_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name or "").lower())


def build_compliance_lookup(audit_df: pd.DataFrame) -> Tuple[Dict[str, Dict], Dict[str, Dict]]:
    by_id: Dict[str, Dict] = {}
    by_name: Dict[str, Dict] = {}

    for _, row in audit_df.iterrows():
        values = {
            check: display_to_binary(resolve_check(row, check).answer_display)
            for check in CHECK_ORDER
        }
        supplier_no = str(row.get("Supplier No.", "")).strip()
        supplier_name = str(row.get("Supplier", "")).strip()
        key = audit_supplier_to_heiko_key(supplier_no)
        if key:
            by_id[key] = values
        if supplier_name:
            by_name[normalize_name(supplier_name)] = values

    return by_id, by_name


# --------------------------------------------------------------------------- #
# XLSX helpers                                                                 #
# --------------------------------------------------------------------------- #
def col_letter(idx: int) -> str:
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def col_index(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


_CELL_RE = re.compile(r'<c\b([^>]*?)(/>|>(.*?)</c>)', re.S)
_STYLE_RE = re.compile(r'\bs="(\d+)"')


def scan_rows_to_update(
    path: str, by_id: Dict[str, Dict], by_name: Dict[str, Dict]
) -> Tuple[Dict[int, Dict[int, object]], int, set]:
    """Return {row_number: {col_index: value}} for all matched rows.

    Uses openpyxl read_only streaming over columns A/B only -> fast, low memory.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    updates: Dict[int, Dict[int, object]] = {}
    matched_ids: set = set()
    data_rows_seen = 0

    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=2, values_only=True), start=1):
        if i < DATA_START_ROW:
            continue
        sid_cell, name_cell = row[0], row[1]
        if sid_cell is not None:
            data_rows_seen += 1

        values = None
        if sid_cell is not None:
            try:
                sid_key = str(int(float(sid_cell)))
            except (TypeError, ValueError):
                sid_key = str(sid_cell).strip()
            values = by_id.get(sid_key)
            if values:
                matched_ids.add(sid_key)
        if values is None and name_cell:
            values = by_name.get(normalize_name(str(name_cell)))
        if values is None:
            continue

        updates[i] = {COL_COMPLIANCE[c]: values[c] for c in COL_COMPLIANCE}

    wb.close()
    return updates, data_rows_seen, matched_ids


def _build_cell_xml(col_idx: int, row_num: int, value, style: Optional[str]) -> str:
    ref = f"{col_letter(col_idx)}{row_num}"
    attrs = f' r="{ref}"'
    if style is not None:
        attrs += f' s="{style}"'
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f"<c{attrs}><v>{int(value)}</v></c>"
    text = str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    attrs += ' t="inlineStr"'
    return f'<c{attrs}><is><t xml:space="preserve">{text}</t></is></c>'


def _rewrite_row(row_open: str, inner: str, row_num: int, col_values: Dict[int, object]) -> str:
    """Rebuild one <row> keeping existing cells, replacing target compliance cells."""
    cells: Dict[int, str] = {}
    styles: Dict[int, Optional[str]] = {}
    for m in _CELL_RE.finditer(inner):
        attrs = m.group(1)
        rm = re.search(r'\br="([A-Z]+)\d+"', attrs)
        if not rm:
            continue
        ci = col_index(rm.group(1))
        cells[ci] = m.group(0)
        sm = _STYLE_RE.search(attrs)
        styles[ci] = sm.group(1) if sm else None

    for ci, value in col_values.items():
        cells[ci] = _build_cell_xml(ci, row_num, value, styles.get(ci))

    ordered = "".join(cells[ci] for ci in sorted(cells))
    return f"{row_open}{ordered}</row>"


_ROW_RE = re.compile(r'(<row\b[^>]*?\br="(\d+)"[^>]*?>)(.*?)</row>', re.S)


def apply_updates_to_sheet_xml(sheet_xml: str, updates: Dict[int, Dict[int, object]]) -> str:
    def repl(m: "re.Match") -> str:
        row_num = int(m.group(2))
        col_values = updates.get(row_num)
        if not col_values:
            return m.group(0)
        return _rewrite_row(m.group(1), m.group(3), row_num, col_values)

    return _ROW_RE.sub(repl, sheet_xml)


def surgical_fill(path: str, updates: Dict[int, Dict[int, object]]) -> None:
    """Rewrite only the compliance cells of matched rows; copy everything else."""
    tmp_path = path + ".tmp"
    with zipfile.ZipFile(path, "r") as zin:
        names = zin.namelist()
        if SHEET_XML not in names:
            raise RuntimeError(f"{SHEET_XML} not found inside workbook")
        sheet_xml = zin.read(SHEET_XML).decode("utf-8")
        new_xml = apply_updates_to_sheet_xml(sheet_xml, updates).encode("utf-8")

        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = new_xml if item.filename == SHEET_XML else zin.read(item.filename)
                zout.writestr(item, data)

    os.replace(tmp_path, path)


def count_data_rows(path: str) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    count = 0
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
        if i >= DATA_START_ROW and row[0] is not None:
            count += 1
    wb.close()
    return count


# --------------------------------------------------------------------------- #
# Main                                                                         #
# --------------------------------------------------------------------------- #
def fill(csv_path: str, xlsx_path: str, make_backup: bool = True) -> None:
    if not os.path.isfile(xlsx_path):
        print(f"Missing file: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    audit_df = pd.read_csv(csv_path, encoding="utf-8-sig")
    by_id, by_name = build_compliance_lookup(audit_df)

    updates, data_rows_seen, matched_ids = scan_rows_to_update(xlsx_path, by_id, by_name)

    if data_rows_seen == 0:
        print(
            f"ERROR: {xlsx_path} appears empty (no Supplier ID in data rows).\n"
            "Restore it from OneDrive version history (or from docs/raw_file.xlsx), "
            "then re-run this script.",
            file=sys.stderr,
        )
        sys.exit(1)

    if make_backup:
        backup = f"{xlsx_path}.bak-{datetime.now():%Y%m%d-%H%M%S}"
        shutil.copy2(xlsx_path, backup)
    else:
        backup = "(skipped)"

    surgical_fill(xlsx_path, updates)

    remaining = count_data_rows(xlsx_path)
    print(f"Audit CSV: {csv_path}")
    print(f"Target xlsx: {xlsx_path}")
    print(f"Audit suppliers in CSV: {len(audit_df)}")
    print(f"Unique supplier IDs mapped: {len(by_id)}")
    print(f"Data rows seen (col A non-empty): {data_rows_seen}")
    print(f"Rows filled: {len(updates)}")
    print(f"Distinct Supplier IDs matched: {len(matched_ids)}")
    print(f"Data rows after write (integrity check): {remaining}")
    print(f"Backup: {backup}")
    print(f"Saved: {xlsx_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fill compliance columns in a Heiko xlsx from a PPWR audit CSV."
    )
    parser.add_argument("--csv", default=AUDIT_CSV, help="Audit results CSV path")
    parser.add_argument("--xlsx", default=HEIKO_XLSX, help="Target Heiko xlsx path")
    parser.add_argument(
        "--no-backup", action="store_true", help="Do not create a .bak copy before writing"
    )
    args = parser.parse_args()
    fill(args.csv, args.xlsx, make_backup=not args.no_backup)


# --------------------------------------------------------------------------- #
# Self-test (no real data needed): build a synthetic workbook and verify.      #
# --------------------------------------------------------------------------- #
def _self_test() -> None:
    from openpyxl import Workbook

    tmp_dir = os.path.join(BASE_DIR, "_selftest_tmp")
    os.makedirs(tmp_dir, exist_ok=True)
    path = os.path.join(tmp_dir, "synthetic.xlsx")

    wb = Workbook()
    ws = wb.active
    ws["A3"] = "Supplier ID"
    ws["B3"] = "Supplier Name"
    ws["F3"] = "Compliance on heavy metals"
    ws["G3"] = "Compliance on PFAS"
    ws["H3"] = "Compliance on SoC"
    ws["I3"] = "Compliance on SVHC"
    # data rows: id, name, plus some pre-existing unrelated columns
    ws["A4"], ws["B4"], ws["J4"], ws["K4"] = 100114879, "ALUBERG SPA", "keepJ", "keepK"
    ws["A5"], ws["B5"], ws["J5"] = 100009453, "CONSTANTIA SAN PROSPERO SRL", "keepJ5"
    ws["A6"], ws["B6"] = 999999999, "UNMATCHED SUPPLIER"
    wb.save(path)

    by_id = {
        "100114879": {"Heavy metals": 1, "PFAS": 1, "SoC": "no response", "SVHC": "no response"},
        "100009453": {"Heavy metals": 1, "PFAS": 1, "SoC": "no response", "SVHC": 0},
    }
    updates, seen, matched = scan_rows_to_update(path, by_id, {})
    assert seen == 3, f"expected 3 data rows, got {seen}"
    assert set(updates) == {4, 5}, f"unexpected rows: {set(updates)}"

    surgical_fill(path, updates)

    wb2 = load_workbook(path, data_only=True)
    ws2 = wb2.active
    assert ws2["F4"].value == 1, ws2["F4"].value
    assert ws2["G4"].value == 1
    assert ws2["H4"].value == "no response", ws2["H4"].value
    assert ws2["I4"].value == "no response"
    assert ws2["I5"].value == 0
    assert ws2["J4"].value == "keepJ", "unrelated column must be preserved"
    assert ws2["K4"].value == "keepK"
    assert ws2["J5"].value == "keepJ5"
    assert ws2["A6"].value == 999999999 and ws2["F6"].value is None, "unmatched row untouched"
    wb2.close()
    shutil.rmtree(tmp_dir, ignore_errors=True)
    print("SELF-TEST PASSED: surgical fill writes correct values and preserves other data.")


if __name__ == "__main__":
    if "--self-test" in sys.argv:
        _self_test()
    else:
        main()
